import pandas as pd
import io
import logging

logger = logging.getLogger(__name__)

def load_master_topics(file):
    """
    Loads topics from the 'Comparison' sheet of the master Excel.
    Handles merged cells by dropping NaN in the Topic column.
    """
    df = pd.read_excel(file, sheet_name='Comparison')
    # Assuming 'Topic' is in a specific column or we find it
    # Based on specs, it's Column B or named 'Topic'
    if 'Topic' not in df.columns:
        # Fallback: maybe the header is on a different row or name differs
        # For now, we assume 'Topic' exists as per spec
        pass
    
    # Filter out empty topics (common with merged cells if logic relies on Column B)
    topics = df['Topic'].dropna().unique().tolist()
    return topics, df

def get_price_duration_columns(file_path):
    """
    Extracts column titles from the 'Price, Duration, Projects' sheet.
    Returns a list of column names from row 1.
    """
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        if "Price, Duration, Projects" not in wb.sheetnames:
            wb.close()
            return []
        
        ws = wb["Price, Duration, Projects"]
        columns = []
        
        # Read row 1 to get column titles (read up to max_column)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                columns.append(str(cell.value).strip())
            else:
                columns.append(None)  # Keep track of empty columns too
        
        wb.close()
        return columns
    except Exception as e:
        # Use basic print if logging not available
        try:
            logger = logging.getLogger(__name__)
            logger.warning(f"Failed to extract columns from 'Price, Duration, Projects' sheet: {e}")
        except:
            print(f"Warning: Failed to extract columns from 'Price, Duration, Projects' sheet: {e}")
        return []

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font
from copy import copy

def copy_cell_style(source_cell, target_cell):
    """
    Copies all formatting from source_cell to target_cell.
    This includes fill, font, border, alignment, number_format, protection, etc.
    """
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.border = copy(source_cell.border)
        target_cell.fill = copy(source_cell.fill)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        target_cell.alignment = copy(source_cell.alignment)

def update_excel_with_analysis(source_file_path, analysis_results, competitor_name, course_name=None, website_link=None, extracted_info=None):
    """
    Updates the Excel file with analysis results using openpyxl to preserve existing comments and formatting.
    Returns the saved workbook as bytes.
    """
    # Load the workbook (data_only=False ensures we keep formulas/comments if any, though we want structure)
    wb = openpyxl.load_workbook(source_file_path)
    
    if "Comparison" not in wb.sheetnames:
        raise ValueError("Sheet 'Comparison' not found in Master Excel.")
        
    ws = wb["Comparison"]
    
    # 1. Determine Target Column (Update existing or Append new)
    target_col_idx = None
    is_new_column = False
    column_header = f"{course_name} by {competitor_name}" if course_name else competitor_name
    
    # Check if header already exists (Row 1)
    for cell in ws[1]:
        if cell.value == column_header:
            target_col_idx = cell.column
            break
            
    # If not found, append to end
    if not target_col_idx:
        target_col_idx = ws.max_column + 1
        is_new_column = True
    
    # 2. Find reference columns to copy styles from
    # For headers: use any competitor column (they all have orange header)
    # For data cells: use column 4 (first competitor column) which has no fill
    header_reference_col_idx = None
    data_reference_col_idx = None
    
    # Find a competitor column (columns 4+) for both header and data cell style
    for check_col in range(max(4, target_col_idx - 1), 2, -1):
        if check_col < target_col_idx:  # Don't check our own column
            test_cell = ws.cell(row=1, column=check_col)
            if test_cell.has_style:
                header_reference_col_idx = check_col
                data_reference_col_idx = check_col
                break
    
    # Fallback: use column 4 if available, otherwise column 3, then column 1
    if data_reference_col_idx is None:
        if ws.max_column >= 4:
            data_reference_col_idx = 4
            if header_reference_col_idx is None:
                header_reference_col_idx = 4
        elif ws.max_column >= 3:
            data_reference_col_idx = 3
            if header_reference_col_idx is None:
                header_reference_col_idx = 3
        else:
            header_reference_col_idx = 1
            data_reference_col_idx = 1
    
    # 3. Write Header (only needed if new, but harmless to overwrite)
    # Row 1 is header
    header_cell = ws.cell(row=1, column=target_col_idx)
    
    # If it's a new column, copy style from reference column before setting value
    if is_new_column and header_reference_col_idx > 0:
        reference_header = ws.cell(row=1, column=header_reference_col_idx)
        copy_cell_style(reference_header, header_cell)
    # If updating existing column, preserve its existing style (don't overwrite)
    
    header_cell.value = column_header
    
    # 4. Iterate rows and populate data
    # We assume Column B (index 2) contains the Topics as per spec
    # starting from row 2
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, min_col=2, max_col=2, values_only=False), start=2):
        cell = row[0]
        topic = str(cell.value).strip() if cell.value else ""
        
        # Stop at TOPIC END
        if topic == "TOPIC END":
            break
            
        if not topic:
            continue
            
        # Get Analysis
        result = analysis_results.get(topic, {})
        decision = result.get('decision', None) # Default None to leave blank if missing? Or "No"? 
        # User spec v1: "No" if missing.
        # But wait, sticking to "No" is safer unless we want to distinguish "Not Analyzed"
        if not decision:
             # If topic exists in analysis_results key but decision is empty -> No
             # If topic is NOT in analysis_results -> No (implied check)
             decision = "No"
             
        reasoning = result.get('reasoning', "No mention found.")
        
        # Write Decision
        target_cell = ws.cell(row=row_idx, column=target_col_idx)
        
        # Copy base style from reference column's corresponding row (only for new columns)
        if is_new_column and data_reference_col_idx > 0:
            reference_cell = ws.cell(row=row_idx, column=data_reference_col_idx)
            copy_cell_style(reference_cell, target_cell)
        # If updating existing column, preserve its existing style (don't overwrite)
        
        target_cell.value = decision
        
        # Apply conditional fill and font colors based on decision value
        # Colors match column 4:
        # Yes: fill=c6efce (green), font=006100 (dark green)
        # No: fill=ffc7ce (red), font=9c0006 (dark red)
        # Unsure/Maybe: fill=f5f19f (yellow), font=61540c (dark yellow/brown)
        decision_lower = str(decision).strip().lower()
        
        # Preserve existing font properties and only update color
        current_font = target_cell.font if target_cell.font else Font()
        
        if decision_lower == "yes":
            target_cell.fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
            target_cell.font = Font(
                name=current_font.name,
                size=current_font.size,
                bold=current_font.bold,
                italic=current_font.italic,
                underline=current_font.underline,
                color="006100"
            )
        elif decision_lower == "no":
            target_cell.fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
            target_cell.font = Font(
                name=current_font.name,
                size=current_font.size,
                bold=current_font.bold,
                italic=current_font.italic,
                underline=current_font.underline,
                color="9c0006"
            )
        elif decision_lower in ["unsure", "maybe"]:
            target_cell.fill = PatternFill(start_color="f5f19f", end_color="f5f19f", fill_type="solid")
            target_cell.font = Font(
                name=current_font.name,
                size=current_font.size,
                bold=current_font.bold,
                italic=current_font.italic,
                underline=current_font.underline,
                color="61540c"
            )
        
        # Write Comment
        if reasoning:
            # Note: openpyxl comments need an Author.
            target_cell.comment = Comment(reasoning, "AgenticAI")
    
    # 5. Update 'Price, Duration, Projects' sheet if data provided
    # Always update if we have course_name or website_link (even if extracted_info is None)
    logger.info(f"Checking Price sheet update: sheet exists={('Price, Duration, Projects' in wb.sheetnames)}, course_name={course_name}, website_link={website_link}, extracted_info={extracted_info}")
    if "Price, Duration, Projects" in wb.sheetnames and (course_name or website_link or extracted_info):
        logger.info("Updating 'Price, Duration, Projects' sheet...")
        ws_pdp = wb["Price, Duration, Projects"]
        
        # Find or create row for this competitor
        target_row_idx = None
        is_new_row = False
        
        # Check if row already exists (match by Provider and Course Name)
        for row_idx, row in enumerate(ws_pdp.iter_rows(min_row=2, values_only=False), start=2):
            provider_cell = row[0]  # Column 1: Provider
            course_cell = row[1]    # Column 2: Course Name
            
            if (provider_cell.value == competitor_name and 
                course_cell.value == course_name):
                target_row_idx = row_idx
                break
        
        # If not found, append new row
        if not target_row_idx:
            target_row_idx = ws_pdp.max_row + 1
            is_new_row = True
        
        # Get column titles to map data dynamically
        columns = get_price_duration_columns(source_file_path)
        
        # Create a mapping of column name to index
        column_map = {}
        for idx, col_name in enumerate(columns, start=1):
            if col_name:
                column_map[col_name.lower()] = idx
        
        # Find a reference row to copy styles from (use row 2 if it exists)
        reference_row_idx = 2 if ws_pdp.max_row >= 2 else None
        
        # Write data based on column mapping
        # Provider (Column 1 or based on "Provider" header)
        provider_col = column_map.get("provider", 1)
        cell = ws_pdp.cell(row=target_row_idx, column=provider_col)
        
        # Copy style from reference row if new row
        if is_new_row and reference_row_idx:
            ref_cell = ws_pdp.cell(row=reference_row_idx, column=provider_col)
            copy_cell_style(ref_cell, cell)
        
        cell.value = competitor_name
        
        # Course Name (Column 2 or based on "Course Name" header)
        if course_name:
            course_col = column_map.get("course name", 2)
            cell = ws_pdp.cell(row=target_row_idx, column=course_col)
            
            # Copy style from reference row if new row
            if is_new_row and reference_row_idx:
                ref_cell = ws_pdp.cell(row=reference_row_idx, column=course_col)
                copy_cell_style(ref_cell, cell)
            
            cell.value = course_name
        
        # Website Link (Column 9 or based on "Website Link" header)
        if website_link:
            website_col = column_map.get("website link", 9)
            cell = ws_pdp.cell(row=target_row_idx, column=website_col)
            
            # Copy style from reference row if new row
            if is_new_row and reference_row_idx:
                ref_cell = ws_pdp.cell(row=reference_row_idx, column=website_col)
                copy_cell_style(ref_cell, cell)
            
            cell.value = website_link
        
        # Write extracted information (Price, Duration, Projects, etc.)
        if extracted_info:
            logger.info(f"Writing extracted_info to row {target_row_idx}: {extracted_info}")
            for col_name, value in extracted_info.items():
                col_idx = column_map.get(col_name.lower())
                if col_idx:
                    cell = ws_pdp.cell(row=target_row_idx, column=col_idx)
                    
                    # Copy style from reference row if new row
                    if is_new_row and reference_row_idx:
                        ref_cell = ws_pdp.cell(row=reference_row_idx, column=col_idx)
                        copy_cell_style(ref_cell, cell)
                    
                    cell.value = value
                    logger.info(f"  Written {col_name} = {value} to column {col_idx}")
                else:
                    logger.warning(f"  Column '{col_name}' not found in column_map. Available columns: {list(column_map.keys())}")
            
    # 6. Save to Bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
