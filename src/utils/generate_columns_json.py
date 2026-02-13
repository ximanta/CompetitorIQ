import pandas as pd
import json
import os
import glob
import openpyxl

# Constants
MASTER_DIR = "src/data/master"

def get_latest_master_file_in_folder(folder_path):
    """Finds the latest version of the master file based on modification time in a specific folder."""
    # Search for ANY .xlsx file in this folder
    search_pattern = os.path.join(folder_path, "*.xlsx")
    files = glob.glob(search_pattern)
    
    # Filter out temp files (starting with ~$)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    
    if not files:
        return None
    return max(files, key=os.path.getmtime)

def generate_columns_for_folder(folder_path):
    file_path = get_latest_master_file_in_folder(folder_path)
    if not file_path:
        print(f"Skipping {folder_path}: No .xlsx file found.")
        return

    print(f"Processing {folder_path}...")
    print(f"  Reading columns from: {os.path.basename(file_path)}")
    
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        
        if "Price, Duration, Projects" not in wb.sheetnames:
            print(f"  ❌ Error: 'Price, Duration, Projects' sheet not found in {os.path.basename(file_path)}.")
            wb.close()
            return
        
        ws = wb["Price, Duration, Projects"]
        columns = []
        
        # Read row 1 to get column titles (read up to max_column)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            if cell.value:
                columns.append(str(cell.value).strip())
            else:
                columns.append(None)  # Keep track of empty columns too
        
        wb.close()
        
        # Filter out None values for cleaner output
        columns_clean = [col for col in columns if col is not None]
        
        output_path = os.path.join(folder_path, "price_duration_columns.json")
        with open(output_path, 'w') as f:
            json.dump(columns_clean, f, indent=4)
            
        print(f"  ✅ Generated price_duration_columns.json with {len(columns_clean)} columns.")
        
    except Exception as e:
        print(f"  ❌ Error processing {os.path.basename(file_path)}: {e}")

def generate_columns():
    # Iterate over all subdirectories in MASTER_DIR
    if not os.path.exists(MASTER_DIR):
        print(f"Error: {MASTER_DIR} does not exist.")
        return

    items = os.listdir(MASTER_DIR)
    subfolders = [os.path.join(MASTER_DIR, item) for item in items if os.path.isdir(os.path.join(MASTER_DIR, item))]
    
    print(f"Found {len(subfolders)} track folders in {MASTER_DIR}")
    
    for folder in subfolders:
        generate_columns_for_folder(folder)

    print("\nGlobal column generation complete.")

if __name__ == "__main__":
    generate_columns()

